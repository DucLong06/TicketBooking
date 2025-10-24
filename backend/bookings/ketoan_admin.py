# backend/bookings/ketoan_admin.py
# THÊM CỘT GHI CHÚ

from django.contrib.admin import AdminSite
from django.contrib import admin
from django.utils.html import format_html
from django.http import HttpResponse
from django.shortcuts import render
from django.urls import path
from django.utils import timezone
from django.db.models import Q, Count, Sum
from datetime import timedelta, datetime
import re

from .models import Booking, SeatReservation
from shows.models import Performance, Show
from venues.models import Seat, PriceCategory, Row

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def parse_invoice_from_notes(notes):
    """Parse thông tin hóa đơn từ notes"""
    if not notes or 'YÊU CẦU XUẤT HOÁ ĐƠN' not in notes:
        return None

    invoice_info = {}
    patterns = {
        'company_name': r'Tên công ty:\s*(.+)',
        'company_address': r'Địa chỉ công ty:\s*(.+)',
        'tax_id': r'Mã số thuế:\s*(.+)',
        'email': r'Email nhận HĐ:\s*(.+)'
    }

    for key, pattern in patterns.items():
        match = re.search(pattern, notes)
        if match:
            invoice_info[key] = match.group(1).strip()

    return invoice_info if invoice_info else None


def extract_clean_notes(notes):
    """Trích xuất ghi chú sạch (bỏ phần hóa đơn)"""
    if not notes:
        return ''

    # Nếu có hóa đơn, lấy phần sau dấu phân cách
    if 'YÊU CẦU XUẤT HOÁ ĐƠN' in notes:
        parts = notes.split('--------------------------------')
        # Lấy phần cuối cùng (ghi chú thực tế)
        clean_notes = parts[-1].strip() if len(parts) > 2 else ''
        return clean_notes

    return notes.strip()


class KeToanAdminSite(AdminSite):
    site_header = '📊 Hệ Thống Kế Toán'
    site_title = 'Kế Toán'
    index_title = 'Dashboard Kế Toán'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('', self.admin_view(self.dashboard_view), name='index'),
            path('export-excel/', self.admin_view(self.export_excel_view), name='export_excel'),
        ]
        return custom_urls + urls

    def dashboard_view(self, request):
        """Dashboard với filter đầy đủ"""

        # Lấy filters
        show_id = request.GET.get('show')
        performance_id = request.GET.get('performance')
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')

        # Query bookings
        bookings_query = Booking.objects.filter(status='paid')

        # Apply filters
        if from_date:
            try:
                from_dt = datetime.strptime(from_date, '%Y-%m-%d')
                bookings_query = bookings_query.filter(paid_at__gte=from_dt)
            except:
                pass
        else:
            # Mặc định 7 ngày gần nhất nếu không có filter
            seven_days_ago = timezone.now() - timedelta(days=7)
            bookings_query = bookings_query.filter(paid_at__gte=seven_days_ago)

        if to_date:
            try:
                to_dt = datetime.strptime(to_date, '%Y-%m-%d')
                to_dt = to_dt.replace(hour=23, minute=59, second=59)
                bookings_query = bookings_query.filter(paid_at__lte=to_dt)
            except:
                pass

        if performance_id:
            bookings_query = bookings_query.filter(performance_id=performance_id)
        elif show_id:
            bookings_query = bookings_query.filter(performance__show_id=show_id)

        recent_bookings = bookings_query.select_related(
            'performance__show'
        ).prefetch_related(
            'seat_reservations__seat__row__section',
            'seat_reservations__seat__price_category',
            'seat_reservations__seat__row__price_category',
            'payments'
        ).order_by('-paid_at')[:100]

        # Xử lý thông tin chi tiết cho từng booking
        for booking in recent_bookings:
            booking.invoice_info = parse_invoice_from_notes(booking.notes)
            booking.clean_notes = extract_clean_notes(booking.notes)

            seats = booking.seat_reservations.all()
            booking.seat_count = len(seats)

            categories = set()
            seat_numbers = []
            for seat in seats:
                pc = seat.seat.effective_price_category
                if pc:
                    categories.add(pc.name)
                seat_numbers.append(seat.seat.full_display_label)

            booking.ticket_classes = ', '.join(sorted(categories)) if categories else '—'
            booking.seat_numbers = ', '.join(seat_numbers)

            # Check payment method (9Pay)
            booking.is_9pay = False
            successful_payment = booking.payments.filter(status='success').first()
            if successful_payment and successful_payment.payment_method == '9pay':
                booking.is_9pay = True

        # ===== STATS THEO FILTER =====
        total_bookings = bookings_query.count()
        total_with_invoice = bookings_query.filter(
            notes__icontains='YÊU CẦU XUẤT HOÁ ĐƠN'
        ).count()

        # Tổng doanh thu
        total_revenue = bookings_query.aggregate(
            total=Sum('final_amount')
        )['total'] or 0

        # Booking qua 9Pay
        total_9pay = bookings_query.filter(
            payments__status='success',
            payments__payment_method='9pay'
        ).distinct().count()

        # ===== VÉ TỒN =====
        upcoming_performances = Performance.objects.filter(
            datetime__gte=timezone.now(),
            status='on_sale'
        ).select_related('show__venue').order_by('datetime')[:10]

        inventory_data = []
        for perf in upcoming_performances:
            venue = perf.show.venue

            total_seats = Seat.objects.filter(
                row__section__venue=venue,
                status='active'
            ).count()

            sold_count = SeatReservation.objects.filter(
                performance=perf,
                booking__status='paid'
            ).count()

            available_count = total_seats - sold_count

            categories_data = []
            used_categories = PriceCategory.objects.filter(
                Q(row__section__venue=venue) |
                Q(seat__row__section__venue=venue)
            ).distinct()

            for pc in used_categories:
                cat_total = Seat.objects.filter(
                    row__section__venue=venue,
                    status='active'
                ).filter(
                    Q(price_category=pc) | Q(row__price_category=pc)
                ).count()

                if cat_total == 0:
                    continue

                cat_sold = SeatReservation.objects.filter(
                    performance=perf,
                    booking__status='paid'
                ).filter(
                    Q(seat__price_category=pc) | Q(seat__row__price_category=pc)
                ).count()

                cat_available = cat_total - cat_sold

                categories_data.append({
                    'name': pc.name,
                    'color': pc.color,
                    'total': cat_total,
                    'sold': cat_sold,
                    'available': cat_available,
                })

            inventory_data.append({
                'performance': perf,
                'total_seats': total_seats,
                'sold_count': sold_count,
                'available_count': available_count,
                'categories': categories_data
            })

        # ===== FILTER OPTIONS =====
        shows = Show.objects.filter(
            performances__datetime__gte=timezone.now()
        ).distinct().order_by('name')

        performances_query = Performance.objects.filter(
            datetime__gte=timezone.now()
        ).select_related('show').order_by('datetime')

        if show_id:
            performances_query = performances_query.filter(show_id=show_id)

        performances = performances_query[:50]

        context = {
            **self.each_context(request),
            'recent_bookings': recent_bookings,
            'total_bookings': total_bookings,
            'total_with_invoice': total_with_invoice,
            'total_revenue': total_revenue,
            'total_9pay': total_9pay,
            'inventory_data': inventory_data,
            'shows': shows,
            'performances': performances,
            'selected_show': int(show_id) if show_id else None,
            'selected_performance': int(performance_id) if performance_id else None,
            'from_date': from_date or '',
            'to_date': to_date or '',
        }

        return render(request, 'admin/ketoan/dashboard.html', context)

    def export_excel_view(self, request):
        """Export Excel với filter"""
        show_id = request.GET.get('show')
        performance_id = request.GET.get('performance')
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')

        queryset = Booking.objects.filter(status='paid')

        if from_date:
            try:
                from_dt = datetime.strptime(from_date, '%Y-%m-%d')
                queryset = queryset.filter(paid_at__gte=from_dt)
            except:
                pass

        if to_date:
            try:
                to_dt = datetime.strptime(to_date, '%Y-%m-%d')
                to_dt = to_dt.replace(hour=23, minute=59, second=59)
                queryset = queryset.filter(paid_at__lte=to_dt)
            except:
                pass

        if performance_id:
            queryset = queryset.filter(performance_id=performance_id)
        elif show_id:
            queryset = queryset.filter(performance__show_id=show_id)

        queryset = queryset.select_related(
            'performance__show'
        ).prefetch_related(
            'seat_reservations__seat__row__section',
            'seat_reservations__seat__price_category',
            'payments'
        ).order_by('-paid_at')

        filename = 'booking_filtered'
        return export_bookings_to_excel(queryset, filename)


ketoan_admin_site = KeToanAdminSite(name='ketoan_admin')


class SeatReservationInlineKeToan(admin.TabularInline):
    model = SeatReservation
    extra = 0
    readonly_fields = ['seat', 'price', 'status']
    can_delete = False
    fields = ['seat', 'price', 'status']


class BookingKeToanAdmin(admin.ModelAdmin):
    list_display = [
        'booking_code',
        'customer_name',
        'customer_phone',
        'performance_info',
        'status_badge',
        'total_seats',
        'amount_display',
        'invoice_badge',
        'paid_at'
    ]
    list_filter = [
        'status',
        'performance__show',
        'paid_at',
        'created_at'
    ]
    search_fields = [
        'booking_code',
        'customer_name',
        'customer_email',
        'customer_phone',
    ]
    readonly_fields = [
        'booking_code',
        'performance',
        'customer_name',
        'customer_email',
        'customer_phone',
        'customer_address',
        'shipping_time',
        'status',
        'total_amount',
        'service_fee',
        'shipping_fee',
        'discount_amount',
        'final_amount',
        'invoice_display',
        'notes_clean',
        'created_at',
        'paid_at'
    ]
    inlines = [SeatReservationInlineKeToan]
    date_hierarchy = 'paid_at'

    fieldsets = (
        ('Thông tin booking', {
            'fields': ('booking_code', 'performance', 'status', 'paid_at')
        }),
        ('Khách hàng', {
            'fields': ('customer_name', 'customer_phone', 'customer_email', 'customer_address', 'shipping_time')
        }),
        ('Thanh toán', {
            'fields': ('total_amount', 'service_fee', 'shipping_fee', 'discount_amount', 'final_amount')
        }),
        ('Hóa đơn & Ghi chú', {
            'fields': ('invoice_display', 'notes_clean')
        })
    )

    actions = ['export_selected']

    def performance_info(self, obj):
        return format_html(
            '<strong>{}</strong><br><small>{}</small>',
            obj.performance.show.name,
            obj.performance.datetime.strftime('%d/%m/%Y %H:%M')
        )
    performance_info.short_description = 'Suất diễn'

    def status_badge(self, obj):
        colors = {'paid': '#4CAF50', 'pending': '#FF9800', 'cancelled': '#F44336'}
        color = colors.get(obj.status, '#999')
        return format_html(
            '<span style="background:{}; color:white; padding:4px 10px; border-radius:4px; font-weight:bold;">{}</span>',
            color, obj.get_status_display()
        )
    status_badge.short_description = 'Trạng thái'

    def total_seats(self, obj):
        return obj.seat_reservations.count()
    total_seats.short_description = 'Số vé'

    def amount_display(self, obj):
        return format_html('<strong>{:,} ₫</strong>', int(obj.final_amount))
    amount_display.short_description = 'Tổng tiền'

    def invoice_badge(self, obj):
        if parse_invoice_from_notes(obj.notes):
            return format_html('<span style="color:#4CAF50; font-weight:bold;">✓ Có</span>')
        return format_html('<span style="color:#ccc;">—</span>')
    invoice_badge.short_description = 'Hóa đơn'

    def invoice_display(self, obj):
        inv = parse_invoice_from_notes(obj.notes)
        if not inv:
            return format_html('<em style="color:#999;">Không yêu cầu xuất hóa đơn</em>')

        html = '<div style="background:#e8f5e9; padding:15px; border-radius:8px; border-left:4px solid #4CAF50;">'
        html += '<h3 style="margin:0 0 10px 0; color:#2e7d32;">📄 Thông tin xuất hóa đơn</h3>'
        html += f'<p><strong>Công ty:</strong> {inv.get("company_name", "")}</p>'
        html += f'<p><strong>Địa chỉ:</strong> {inv.get("company_address", "")}</p>'
        html += f'<p><strong>MST:</strong> {inv.get("tax_id", "")}</p>'
        html += f'<p><strong>Email:</strong> {inv.get("email", "")}</p>'
        html += '</div>'
        return format_html(html)
    invoice_display.short_description = 'Thông tin hóa đơn'

    def notes_clean(self, obj):
        notes = extract_clean_notes(obj.notes)
        if not notes:
            return format_html('<em style="color:#999;">Không có ghi chú</em>')

        return format_html('<div style="background:#fff9c4; padding:10px; border-radius:4px;">{}</div>', notes)
    notes_clean.short_description = 'Ghi chú'

    def export_selected(self, request, queryset):
        return export_bookings_to_excel(queryset, 'booking_da_chon')
    export_selected.short_description = '📊 Xuất Excel'

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


def export_bookings_to_excel(queryset, filename_prefix):
    """Export bookings - 14 cột (thêm cột Ghi chú)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bookings"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = [
        'Mã booking', 'Tên KH', 'SĐT', 'Địa chỉ',
        'Suất diễn', 'Ngày giờ', 'Hạng vé', 'Số ghế', 'SL vé',
        'Tổng tiền', '9Pay', 'Thông tin hóa đơn', 'Ghi chú', 'Ngày thanh toán'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    row_num = 2
    for booking in queryset:
        inv = parse_invoice_from_notes(booking.notes)
        clean_notes = extract_clean_notes(booking.notes)

        seats = booking.seat_reservations.all()
        categories = set()
        seat_nums = []

        for s in seats:
            pc = s.seat.effective_price_category
            if pc:
                categories.add(pc.name)
            seat_nums.append(s.seat.full_display_label)

        hang_ve = ', '.join(sorted(categories)) if categories else '—'
        so_ghe = ', '.join(seat_nums)

        invoice_text = ''
        if inv:
            invoice_text = f"CT: {inv.get('company_name', '')}\nMST: {inv.get('tax_id', '')}\nEmail: {inv.get('email', '')}"
        else:
            invoice_text = 'Không có'

        # Check 9Pay
        is_9pay = 'Không'
        successful_payment = booking.payments.filter(status='success').first()
        if successful_payment and successful_payment.payment_method == '9pay':
            is_9pay = 'Có'

        row_data = [
            booking.booking_code,
            booking.customer_name,
            booking.customer_phone,
            booking.customer_address,
            booking.performance.show.name,
            booking.performance.datetime.strftime('%d/%m/%Y %H:%M'),
            hang_ve,
            so_ghe,
            len(seats),
            float(booking.final_amount),
            is_9pay,
            invoice_text,
            clean_notes or 'Không có',
            booking.paid_at.strftime('%d/%m/%Y %H:%M') if booking.paid_at else ''
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border

            if col_num == 10:
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        row_num += 1

    for col in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{timestamp}.xlsx"'
    wb.save(response)
    return response


ketoan_admin_site.register(Booking, BookingKeToanAdmin)
